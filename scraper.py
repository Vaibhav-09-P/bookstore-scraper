#!/usr/bin/env python3
"""
Books to Scrape — portfolio web scraper.
Collects book data across all catalog pages, applies filters, and exports to Excel.
"""

from __future__ import annotations

import argparse
import re
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import config

RATING_WORDS = {
    "one": 1,
    "two": 2,
    "three": 3,
    "four": 4,
    "five": 5,
}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
}


@dataclass
class Book:
    title: str = ""
    price: float | None = None
    price_display: str = ""
    rating: int | None = None
    availability: str = ""
    category: str = ""
    url: str = ""


@dataclass
class ScrapeStats:
    total_scanned: int = 0
    total_exported: int = 0
    skipped_pages: list[str] = field(default_factory=list)
    skipped_items: list[str] = field(default_factory=list)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Scrape books.toscrape.com and export filtered results to Excel."
    )
    parser.add_argument(
        "--min-rating",
        type=int,
        default=config.MIN_RATING,
        help="Minimum star rating (1-5). Use 0 to disable.",
    )
    parser.add_argument(
        "--max-price",
        type=float,
        default=config.MAX_PRICE,
        help="Maximum price in GBP. Use 0 to disable.",
    )
    parser.add_argument(
        "--category",
        type=str,
        default=config.CATEGORY,
        help='Category filter (partial match), e.g. "Mystery".',
    )
    parser.add_argument(
        "--max-results",
        type=int,
        default=config.MAX_RESULTS,
        help="Maximum books to export. Use 0 for no limit.",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=config.OUTPUT_FILE,
        help="Output .xlsx file path.",
    )
    parser.add_argument(
        "--base-url",
        type=str,
        default=config.BASE_URL,
        help="Site base URL.",
    )
    return parser.parse_args()


def normalize_filters(args: argparse.Namespace) -> dict[str, Any]:
    min_rating = args.min_rating if args.min_rating and args.min_rating > 0 else None
    max_price = args.max_price if args.max_price and args.max_price > 0 else None
    category = args.category.strip() if args.category else None
    max_results = args.max_results if args.max_results and args.max_results > 0 else None
    return {
        "min_rating": min_rating,
        "max_price": max_price,
        "category": category,
        "max_results": max_results,
        "output": args.output,
        "base_url": args.base_url.rstrip("/") + "/",
    }


def resolve_book_url(href: str, base_url: str) -> str:
    """Build an absolute book URL (listing pages use mixed relative paths)."""
    href = href.lstrip("/")
    if href.startswith("catalogue/"):
        return urljoin(base_url, href)
    return urljoin(base_url, f"catalogue/{href}")


def fetch_page(
    session: requests.Session,
    url: str,
    stats: ScrapeStats,
    *,
    log_label: str = "page",
) -> BeautifulSoup | None:
    for attempt in range(1, config.MAX_RETRIES + 1):
        try:
            response = session.get(url, headers=HEADERS, timeout=config.REQUEST_TIMEOUT)
            response.raise_for_status()
            response.encoding = response.apparent_encoding or response.encoding
            return BeautifulSoup(response.text, "html.parser")
        except requests.RequestException as exc:
            if attempt == config.MAX_RETRIES:
                entry = f"{url} ({exc})"
                if log_label == "item":
                    stats.skipped_items.append(entry)
                    print(f"  [!] Skipping item after {config.MAX_RETRIES} retries: {url}")
                else:
                    stats.skipped_pages.append(entry)
                    print(f"  [!] Skipping page after {config.MAX_RETRIES} retries: {url}")
                return None
            time.sleep(1.5 * attempt)
    return None


def parse_total_books(soup: BeautifulSoup) -> int | None:
    text = soup.get_text(" ", strip=True)
    match = re.search(r"(\d+)\s+results", text, re.IGNORECASE)
    return int(match.group(1)) if match else None


def parse_rating(element) -> int | None:
    if element is None:
        return None
    classes = element.get("class", [])
    for name in classes:
        if name != "star-rating" and name.lower() in RATING_WORDS:
            return RATING_WORDS[name.lower()]
    return None


def parse_price(text: str) -> tuple[float | None, str]:
    if not text:
        return None, ""
    cleaned = text.strip()
    match = re.search(r"[\d,.]+", cleaned.replace(",", ""))
    if not match:
        return None, cleaned
    try:
        return float(match.group()), cleaned
    except ValueError:
        return None, cleaned


def parse_availability(text: str) -> str:
    if not text:
        return ""
    lower = text.lower()
    if "in stock" in lower:
        return "In stock"
    if "out of stock" in lower:
        return "Out of stock"
    return text.strip()


def parse_listing_book(article, base_url: str) -> Book:
    book = Book()
    title_tag = article.select_one("h3 a")
    if title_tag:
        book.title = title_tag.get("title") or title_tag.get_text(strip=True)
        href = title_tag.get("href", "")
        book.url = resolve_book_url(href, base_url)

    price_tag = article.select_one("p.price_color")
    if price_tag:
        book.price, book.price_display = parse_price(price_tag.get_text())

    book.rating = parse_rating(article.select_one("p.star-rating"))

    avail_tag = article.select_one("p.instock.availability, p.availability")
    if avail_tag:
        book.availability = parse_availability(avail_tag.get_text())

    return book


def fetch_category(session: requests.Session, book_url: str, stats: ScrapeStats) -> str:
    soup = fetch_page(session, book_url, stats, log_label="item")
    if soup is None:
        return ""
    crumbs = soup.select("ul.breadcrumb li:not(.active) a")
    if len(crumbs) >= 2:
        return crumbs[-1].get_text(strip=True)
    if crumbs:
        return crumbs[-1].get_text(strip=True)
    return ""


def passes_filters(book: Book, filters: dict[str, Any]) -> bool:
    if filters["min_rating"] is not None:
        if book.rating is None or book.rating < filters["min_rating"]:
            return False
    if filters["max_price"] is not None:
        if book.price is None or book.price > filters["max_price"]:
            return False
    if filters["category"]:
        if not book.category or filters["category"].lower() not in book.category.lower():
            return False
    return True


def listing_passes_quick_filters(book: Book, filters: dict[str, Any]) -> bool:
    """Skip detail-page fetch when listing data already fails rating/price filters."""
    if filters["min_rating"] is not None:
        if book.rating is None or book.rating < filters["min_rating"]:
            return False
    if filters["max_price"] is not None:
        if book.price is None or book.price > filters["max_price"]:
            return False
    return True


def get_next_page_url(soup: BeautifulSoup, current_url: str) -> str | None:
    next_link = soup.select_one("li.next a")
    if not next_link:
        return None
    return urljoin(current_url, next_link["href"])


def update_progress(scanned: int, total: int | None) -> None:
    if total:
        line = f"\rScraped {scanned} / {total} books"
    else:
        line = f"\rScraped {scanned} books"
    sys.stdout.write(line)
    sys.stdout.flush()


def export_excel(
    books: list[Book],
    filters: dict[str, Any],
    stats: ScrapeStats,
    output_path: str,
) -> None:
    wb = Workbook()

    # --- Results sheet ---
    ws = wb.active
    ws.title = "Results"
    columns = ["Title", "Price", "Rating", "Availability", "Category", "URL"]
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for book in books:
        price_cell = book.price_display if book.price_display else book.price
        ws.append(
            [
                book.title,
                price_cell if price_cell is not None else "",
                book.rating if book.rating is not None else "",
                book.availability,
                book.category,
                book.url,
            ]
        )

    for idx, _ in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 18
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["F"].width = 55

    # --- Summary sheet ---
    summary = wb.create_sheet("Scrape Summary")
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 50

    filter_lines = []
    if filters["min_rating"] is not None:
        filter_lines.append(f"Minimum rating: {filters['min_rating']}+ stars")
    else:
        filter_lines.append("Minimum rating: none")
    if filters["max_price"] is not None:
        filter_lines.append(f"Maximum price: £{filters['max_price']:.2f}")
    else:
        filter_lines.append("Maximum price: none")
    if filters["category"]:
        filter_lines.append(f"Category: {filters['category']}")
    else:
        filter_lines.append("Category: all")
    if filters["max_results"] is not None:
        filter_lines.append(f"Max results: {filters['max_results']}")
    else:
        filter_lines.append("Max results: no limit")

    rows = [
        ("Scrape date & time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total books scanned", stats.total_scanned),
        ("Filters applied", "\n".join(filter_lines)),
        ("Total results exported", stats.total_exported),
        ("Pages skipped", len(stats.skipped_pages)),
        ("Items skipped", len(stats.skipped_items)),
    ]
    for label, value in rows:
        summary.append([label, value])

    if stats.skipped_pages:
        summary.append([])
        summary.append(["Skipped pages", ""])
        for entry in stats.skipped_pages:
            summary.append(["", entry])

    if stats.skipped_items:
        summary.append([])
        summary.append(["Skipped items", ""])
        for entry in stats.skipped_items[:50]:
            summary.append(["", entry])
        if len(stats.skipped_items) > 50:
            summary.append(["", f"... and {len(stats.skipped_items) - 50} more"])

    wb.save(output_path)


def scrape(filters: dict[str, Any]) -> tuple[list[Book], ScrapeStats]:
    session = requests.Session()
    stats = ScrapeStats()
    results: list[Book] = []
    page_url: str | None = filters["base_url"]
    total_books: int | None = None

    while page_url:
        soup = fetch_page(session, page_url, stats)
        if soup is None:
            page_url = None
            continue

        if total_books is None:
            total_books = parse_total_books(soup)
            if total_books:
                print(f"Found {total_books} books across the catalog.\n")

        articles = soup.select("article.product_pod")
        for article in articles:
            book = parse_listing_book(article, filters["base_url"])
            stats.total_scanned += 1
            update_progress(stats.total_scanned, total_books)

            if not listing_passes_quick_filters(book, filters):
                continue

            book.category = fetch_category(session, book.url, stats)

            if not passes_filters(book, filters):
                continue

            results.append(book)
            if filters["max_results"] and len(results) >= filters["max_results"]:
                print()
                return results, stats

        page_url = get_next_page_url(soup, page_url)

    print()
    return results, stats


def print_summary(books: list[Book], filters: dict[str, Any], stats: ScrapeStats, output_path: str) -> None:
    stats.total_exported = len(books)
    print(f"Found {stats.total_exported} books matching your filters.")
    print(f"Exported to: {output_path}")
    if stats.skipped_pages:
        print(f"Skipped {len(stats.skipped_pages)} page(s) — see Excel summary for details.")
    if stats.skipped_items:
        print(f"Skipped {len(stats.skipped_items)} item(s) — see Excel summary for details.")


def main() -> int:
    args = parse_args()
    filters = normalize_filters(args)

    print("Books to Scrape — starting scrape\n")
    print("Active filters:")
    print(f"  Min rating: {filters['min_rating'] or 'none'}")
    print(f"  Max price:  £{filters['max_price']:.2f}" if filters["max_price"] else "  Max price:  none")
    print(f"  Category:   {filters['category'] or 'all'}")
    print(f"  Max export: {filters['max_results'] or 'no limit'}")
    print()

    books, stats = scrape(filters)
    stats.total_exported = len(books)
    export_excel(books, filters, stats, filters["output"])
    print_summary(books, filters, stats, filters["output"])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
